#!/usr/bin/env python3
"""Collect AIT JSON reports and append key metrics to CSV/XLSX."""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

DEFAULT_FIELDS: List[str] = [
    "timestamp",
    "model",
    "protocol",
    "base_url",
    "target_ip",
    "total_requests",
    "concurrency",
    "is_stream",
    "is_thinking",
    "success_rate",
    "error_rate",
    "avg_total_time",
    "avg_ttft",
    "avg_tpot",
    "avg_tps",
    "avg_total_throughput_tps",
    "system_output_tps",
    "system_total_tps",
    "total_output_tokens",
]


class ReportCollector:
    """Incrementally ingests AIT JSON report files into a history table."""

    def __init__(self, output_file: Path, state_file: Path, fields: Sequence[str] | None = None) -> None:
        self.output_file = output_file.resolve()
        self.state_file = state_file.resolve()
        self.fields = list(fields or DEFAULT_FIELDS)
        self._processed = set(self._load_state().get("processed_keys", []))

    def _load_state(self) -> Dict[str, List[str]]:
        if not self.state_file.exists():
            return {"processed_keys": []}
        try:
            content = json.loads(self.state_file.read_text(encoding="utf-8"))
            keys = content.get("processed_keys", [])
            if isinstance(keys, list):
                return {"processed_keys": [str(x) for x in keys]}
        except Exception:
            pass
        return {"processed_keys": []}

    def _save_state(self) -> None:
        payload = {"processed_keys": sorted(self._processed)}
        self.state_file.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )

    @staticmethod
    def _row_key(report_file: Path, report_timestamp: str, model_data: Dict[str, object]) -> str:
        model = str(model_data.get("model", ""))
        protocol = str(model_data.get("protocol", ""))
        total_requests = str(model_data.get("total_requests", ""))
        concurrency = str(model_data.get("concurrency", ""))
        return (
            f"{report_file.resolve()}|{report_timestamp}|{model}|"
            f"{protocol}|{total_requests}|{concurrency}"
        )

    def _extract_rows(self, report_file: Path) -> List[Dict[str, object]]:
        try:
            content = json.loads(report_file.read_text(encoding="utf-8"))
        except Exception as exc:
            print(f"[WARN] Failed to parse {report_file}: {exc}", file=sys.stderr)
            return []

        if not isinstance(content, dict):
            return []
        if content.get("report_type") != "ait_benchmark_report":
            return []

        report_timestamp = str(content.get("timestamp", ""))
        models = content.get("models", [])
        if not isinstance(models, list):
            return []

        rows: List[Dict[str, object]] = []
        for model_data in models:
            if not isinstance(model_data, dict):
                continue
            key = self._row_key(report_file, report_timestamp, model_data)
            if key in self._processed:
                continue

            row: Dict[str, object] = {}
            for field in self.fields:
                if field == "report_file":
                    row[field] = str(report_file.resolve())
                    continue
                if field == "report_timestamp":
                    row[field] = report_timestamp
                    continue
                row[field] = model_data.get(field, "")

            rows.append(row)
            self._processed.add(key)
        return rows

    def _append_rows_csv(self, rows: Sequence[Dict[str, object]]) -> None:
        self.output_file.parent.mkdir(parents=True, exist_ok=True)
        needs_header = not self.output_file.exists() or self.output_file.stat().st_size == 0
        with self.output_file.open("a", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(self.fields), extrasaction="ignore")
            if needs_header:
                writer.writeheader()
            for row in rows:
                writer.writerow(row)

    def _append_rows_xlsx(self, rows: Sequence[Dict[str, object]]) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except Exception as exc:
            raise RuntimeError(
                "Writing .xlsx requires openpyxl. Install with: pip install openpyxl"
            ) from exc

        self.output_file.parent.mkdir(parents=True, exist_ok=True)
        if self.output_file.exists():
            wb = load_workbook(self.output_file)
            ws = wb.active
            if ws.max_row == 0:
                ws.append(list(self.fields))
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "AIT Benchmark"
            ws.append(list(self.fields))

        for row in rows:
            ws.append([row.get(field, "") for field in self.fields])
        wb.save(self.output_file)

    def _append_rows(self, rows: Sequence[Dict[str, object]]) -> None:
        if not rows:
            return
        if self.output_file.suffix.lower() == ".xlsx":
            self._append_rows_xlsx(rows)
        else:
            self._append_rows_csv(rows)

    def ingest_report_files(self, report_files: Sequence[Path]) -> int:
        appended = 0
        all_rows: List[Dict[str, object]] = []
        for report_file in sorted(set(Path(p).resolve() for p in report_files)):
            all_rows.extend(self._extract_rows(report_file))

        if all_rows:
            self._append_rows(all_rows)
            self._save_state()
            appended = len(all_rows)
        return appended

    def ingest_reports_in_dir(self, watch_dir: Path, pattern: str = "ait-report-*.json") -> int:
        files = sorted(watch_dir.glob(pattern), key=lambda p: p.stat().st_mtime)
        return self.ingest_report_files(files)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract key metrics from AIT JSON reports and append them to CSV/XLSX."
    )
    parser.add_argument("--watch-dir", default=".", help="Directory to scan for AIT JSON reports.")
    parser.add_argument("--pattern", default="ait-report-*.json", help="Report filename pattern.")
    parser.add_argument("--output", default="benchmark_history.csv", help="Output file path.")
    parser.add_argument(
        "--state-file",
        default=".ait_report_ingest_state.json",
        help="State file used for de-duplication across restarts.",
    )
    parser.add_argument("--poll-interval", type=float, default=2.0, help="Polling interval seconds.")
    parser.add_argument("--once", action="store_true", help="Process once and exit.")
    parser.add_argument(
        "--fields",
        default=",".join(DEFAULT_FIELDS),
        help="Comma-separated fields to export.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    watch_dir = Path(args.watch_dir).resolve()
    if not watch_dir.exists() or not watch_dir.is_dir():
        print(f"[ERROR] watch-dir is not a valid directory: {watch_dir}", file=sys.stderr)
        return 2

    fields = [x.strip() for x in args.fields.split(",") if x.strip()]
    collector = ReportCollector(
        output_file=Path(args.output),
        state_file=Path(args.state_file),
        fields=fields,
    )

    print(f"[INFO] Watching directory: {watch_dir}")
    print(f"[INFO] Report pattern: {args.pattern}")
    print(f"[INFO] Output file: {collector.output_file}")
    print(f"[INFO] Fields: {', '.join(fields)}")
    print(f"[INFO] De-dup state file: {collector.state_file}")

    while True:
        appended = collector.ingest_reports_in_dir(watch_dir=watch_dir, pattern=args.pattern)
        if appended:
            print(f"[INFO] Appended {appended} row(s) into {collector.output_file}")
        if args.once:
            break
        time.sleep(args.poll_interval)

    print("[INFO] Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
